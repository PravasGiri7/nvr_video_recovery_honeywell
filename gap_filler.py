import os
import re
import sys
import shutil
import struct
import argparse
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Baseline folder (the camera channel with gaps) ──────────
BASELINE_DIR = Path(r"/PATH/TO/BASELINE/DIR")

# ── Fill folder (chunks that may plug the gaps) ──────────────
FILL_DIR = Path(r"/PATH/TO/FILL/DIR")

# ── Output folder (merged result; created automatically) ─────
OUTPUT_DIR = Path(r"/PATH/TO/OUTPUT/DIR")

# ── Gap threshold: gaps larger than this are flagged / filled ─
GAP_THRESHOLD_SECONDS = 5.0

MATCH_TOLERANCE_SECONDS = 3.0

PARTIAL_FILL_MIN_COVERAGE = 0.70   

DAT_EXTENSIONS = {".dat"}

DAT_MARKERS = {
    b'\x80\x01\x00': "80 01 00 (non-IDR frame)",
    b'\x81\x01\x00': "81 01 00 (IDR keyframe)",
}

TS_MIN_US = 1_577_836_800 * 1_000_000   
TS_MAX_US = 2_208_988_800 * 1_000_000  


def _find_all(data: bytes, pattern: bytes) -> list[int]:
    offsets, start = [], 0
    while True:
        pos = data.find(pattern, start)
        if pos == -1:
            break
        offsets.append(pos)
        start = pos + 1
    return offsets


def _read_ts_us(data: bytes, offset: int) -> int | None:
    if offset < 0 or offset + 8 > len(data):
        return None
    return struct.unpack_from('<Q', data, offset)[0]


def _is_valid(ts_us: int) -> bool:
    return TS_MIN_US <= ts_us <= TS_MAX_US


def _known_markers(data: bytes) -> list[dict]:
    results = []
    for marker in DAT_MARKERS:
        for pos in _find_all(data, marker):
            ts = _read_ts_us(data, pos + 11)
            if ts and _is_valid(ts):
                results.append({'offset': pos, 'ts_us': ts})
    return sorted(results, key=lambda x: x['offset'])



def extract_timestamps(path: Path) -> tuple[datetime | None, datetime | None, str]:
    try:
        data = path.read_bytes()
        records = _known_markers(data)
        method = "known_markers"
        if not records:
            method = "alt_anchors"
        if not records:
            return None, None, "none"

        ts_values = [r['ts_us'] for r in records]
        first = datetime.fromtimestamp(min(ts_values) / 1e6, tz=timezone.utc)
        last  = datetime.fromtimestamp(max(ts_values) / 1e6, tz=timezone.utc)
        return first, last, method
    except Exception as e:
        return None, None, f"error:{e}"


@dataclass
class Chunk:
    path:       Path
    first_ts:   datetime | None
    last_ts:    datetime | None
    ts_method:  str
    ts_ok:      bool
    source:     str = "baseline"   

    @property
    def duration_s(self) -> float:
        if self.ts_ok:
            return (self.last_ts - self.first_ts).total_seconds()
        return 0.0


@dataclass
class Gap:
    after:      Chunk
    before:     Chunk
    gap_s:      float
    fillers:    list[Chunk] = field(default_factory=list)
    partial:    bool = False   
    coverage:   float = 0.0   

def scan_folder(folder: Path, source_label: str) -> list[Chunk]:
    files = sorted(
        p for p in folder.iterdir()
        if p.is_file() and p.suffix.lower() in DAT_EXTENSIONS
    )
    if not files:
        print(f"  [WARN] No .dat files in: {folder}")
        return []

    chunks = []
    for p in files:
        first, last, method = extract_timestamps(p)
        ts_ok = first is not None and last is not None
        chunks.append(Chunk(
            path=p, first_ts=first, last_ts=last,
            ts_method=method, ts_ok=ts_ok, source=source_label,
        ))
    return chunks


def detect_gaps(chunks: list[Chunk], threshold_s: float) -> list[Gap]:
    gaps = []
    for i in range(len(chunks) - 1):
        c, n = chunks[i], chunks[i + 1]
        if not c.ts_ok or not n.ts_ok:
            continue
        gap_s = (n.first_ts - c.last_ts).total_seconds()
        if gap_s > threshold_s:
            gaps.append(Gap(after=c, before=n, gap_s=gap_s))
    return gaps


def _td(seconds: float) -> timedelta:
    return timedelta(seconds=seconds)


def _chain_coverage(chain: list[Chunk], gap_open: datetime,
                    gap_close: datetime) -> float:
    gap_s = (gap_close - gap_open).total_seconds()
    if gap_s <= 0:
        return 1.0
    covered_start = max(chain[0].first_ts,  gap_open)
    covered_end   = min(chain[-1].last_ts,  gap_close)
    covered_s     = (covered_end - covered_start).total_seconds()
    return max(0.0, covered_s / gap_s)


def _chain_stats(chain: list[Chunk], gap_open: datetime,
                 gap_close: datetime, tol_s: float) -> dict:
    first = chain[0]
    last  = chain[-1]

    coverage    = _chain_coverage(chain, gap_open, gap_close)
    start_error = abs((first.first_ts - gap_open).total_seconds())
    end_error   = abs((last.last_ts   - gap_close).total_seconds())

    missing_before = max(0.0, (first.first_ts - gap_open).total_seconds())
    missing_after  = max(0.0, (gap_close - last.last_ts).total_seconds())

    overhang_before = max(0.0, (gap_open  - first.first_ts).total_seconds())
    overhang_after  = max(0.0, (last.last_ts - gap_close).total_seconds())

    internal_gap_s = 0.0
    total_overlap_s = 0.0
    for i in range(len(chain) - 1):
        delta = (chain[i + 1].first_ts - chain[i].last_ts).total_seconds()
        if delta >= 0:
            internal_gap_s = max(internal_gap_s, delta)
        else:
            total_overlap_s += abs(delta)

    n_chunks = len(chain)


    full = (
        missing_before  <= tol_s
        and missing_after   <= tol_s
        and overhang_before <= tol_s
        and overhang_after  <= tol_s
        and coverage >= 0.95
    )

    return {
        "coverage":       coverage,
        "full":           full,
        "start_error":    start_error,
        "end_error":      end_error,
        "missing_before": missing_before,
        "missing_after":  missing_after,
        "overhang_before": overhang_before,
        "overhang_after":  overhang_after,
        "internal_gap_s": internal_gap_s,
        "total_overlap_s": total_overlap_s,
        "n_chunks":       n_chunks,
    }


def _chain_sort_key(chain: list[Chunk], gap_open: datetime,
                    gap_close: datetime, tol_s: float):
    s = _chain_stats(chain, gap_open, gap_close, tol_s)

    missing_total  = s["missing_before"] + s["missing_after"]
    overhang_total = s["overhang_before"] + s["overhang_after"]
    edge_error     = s["start_error"] + s["end_error"]

    return (
        0 if s["full"] else 1,          # (1) full chains first
        round(overhang_total, 3),       # (2) LEAST baseline overlap — highest priority penalty
        round(missing_total, 3),        # (3) less missing gap is better
        round(s["internal_gap_s"], 3),  # (4) penalise fragmented chains
        round(edge_error, 3),           # (5) closer edge match is better
        s["n_chunks"],                  # (6) fewer chunks → simpler / more precise
        round(s["total_overlap_s"], 3), # (7) less internal overlap is cleaner
        -round(s["coverage"], 6),       # (8) coverage tie-break
        chain[0].first_ts,             # (9) deterministic tie-break
        chain[0].path.name,
    )


def _prune_candidates(
    candidates: list[Chunk],
    gap_open: datetime,
    gap_close: datetime,
    tol_s: float,
) -> list[Chunk]:

    pruned = []
    for c in candidates:
        if not c.ts_ok:
            continue
        if c.last_ts  <= gap_open  - _td(tol_s):
            continue
        if c.first_ts >= gap_close + _td(tol_s):
            continue
        if c.duration_s <= 0:
            continue
        pruned.append(c)
    return pruned


def _can_follow(prev: Chunk, nxt: Chunk, tol_s: float) -> bool:

    if nxt.first_ts <= prev.first_ts:
        return False                    # must be strictly later start

    boundary_delta = (nxt.first_ts - prev.last_ts).total_seconds()
    if abs(boundary_delta) > tol_s:
        return False                    # too big a gap or too much overlap

    if nxt.last_ts <= prev.last_ts:
        return False                    # nxt is entirely inside prev → skip

    return True


def _find_chain(
    fill_chunks: list[Chunk],
    gap_open: datetime,
    gap_close: datetime,
    tol_s: float,
    min_coverage: float,
) -> tuple[list[Chunk] | None, bool]:

    candidates = _prune_candidates(fill_chunks, gap_open, gap_close, tol_s)
    candidates.sort(key=lambda c: (c.first_ts, c.last_ts, c.path.name))

    if not candidates:
        return None, False

    all_chains: list[list[Chunk]] = []

    def maybe_add(chain: list[Chunk]) -> None:
        if not chain:
            return
        stats = _chain_stats(chain, gap_open, gap_close, tol_s)
        if stats["full"] or stats["coverage"] >= min_coverage:
            all_chains.append(chain)

    def dfs(chain: list[Chunk], used_idx: set[int]) -> None:
        maybe_add(chain)

        last = chain[-1]

        # ── Early exit: chain already reaches (or passes) gap_close ──
        if last.last_ts >= gap_close - _td(tol_s):
            return

        for i, nxt in enumerate(candidates):
            if i in used_idx:
                continue

            # Skip anything that starts so late it can't join the chain
            boundary = (nxt.first_ts - last.last_ts).total_seconds()
            if boundary > tol_s:
                # candidates are sorted by first_ts; everything further
                # is even later, so we can break entirely.
                break

            if _can_follow(last, nxt, tol_s):
                dfs(chain + [nxt], used_idx | {i})

    # Try every candidate as a chain start
    for i, c in enumerate(candidates):
        dfs([c], {i})

    if not all_chains:
        return None, False

    # De-duplicate by path sequence
    unique: dict[tuple[Path, ...], list[Chunk]] = {}
    for chain in all_chains:
        key = tuple(c.path for c in chain)
        unique[key] = chain

    best = sorted(
        unique.values(),
        key=lambda chain: _chain_sort_key(chain, gap_open, gap_close, tol_s),
    )[0]

    best_stats  = _chain_stats(best, gap_open, gap_close, tol_s)
    is_partial  = not best_stats["full"]

    return best, is_partial


def match_fillers(gaps: list[Gap], fill_chunks: list[Chunk],
                  tol_s: float = MATCH_TOLERANCE_SECONDS,
                  min_coverage: float = PARTIAL_FILL_MIN_COVERAGE):
 
    used_paths: set[Path] = set()
    sorted_gaps = sorted(gaps, key=lambda g: g.gap_s, reverse=True)

    for gap in sorted_gaps:
        available = [fc for fc in fill_chunks if fc.path not in used_paths]
        chain, partial = _find_chain(
            available, gap.after.last_ts, gap.before.first_ts,
            tol_s=tol_s, min_coverage=min_coverage,
        )
        if chain:
            gap.fillers  = chain
            gap.partial  = partial
            gap.coverage = _chain_coverage(chain, gap.after.last_ts,
                                           gap.before.first_ts)
            for fc in chain:
                used_paths.add(fc.path)


# MERGE & COPY
def build_merged_sequence(baseline: list[Chunk], gaps: list[Gap]) -> list[Chunk]:
    insert_after: dict[Path, list[Chunk]] = {}
    for gap in gaps:
        if gap.fillers:
            sorted_fillers = sorted(gap.fillers, key=lambda c: c.first_ts)
            insert_after[gap.after.path] = sorted_fillers

    merged: list[Chunk] = []
    for bc in baseline:
        merged.append(bc)
        if bc.path in insert_after:
            merged.extend(insert_after[bc.path])
    return merged


def copy_merged(merged: list[Chunk], output_dir: Path, dry_run: bool):
    output_dir.mkdir(parents=True, exist_ok=True)
    width = len(str(len(merged)))

    print(f"\n  {'DRY RUN — ' if dry_run else ''}Copying {len(merged)} files → {output_dir}\n")

    for idx, chunk in enumerate(merged, start=1):
        new_name = f"{idx:0{width}}_{chunk.path.stem}_{chunk.source}{chunk.path.suffix}"
        dest = output_dir / new_name

        status = "SKIP (exists)" if dest.exists() else ("COPY" if not dry_run else "DRY")
        print(f"  [{idx:>{width}}] {chunk.path.name:50s} → {new_name}  [{status}]")

        if not dry_run and not dest.exists():
            shutil.copy2(chunk.path, dest)

    print()

def fmt_ts(dt: datetime | None) -> str:
    return dt.strftime("%Y/%m/%d %H:%M:%S UTC") if dt else "N/A"


def fmt_dur(s: float) -> str:
    h, rem = divmod(abs(s), 3600)
    m, sec = divmod(rem, 60)
    if h:   return f"{int(h)}h {int(m)}m {sec:.1f}s"
    if m:   return f"{int(m)}m {sec:.1f}s"
    return f"{sec:.3f}s"


def print_chunk_table(chunks: list[Chunk], title: str):
    print(f"\n  ── {title} ({'%d chunk(s)' % len(chunks)}) ──")
    col = max((len(c.path.name) for c in chunks), default=10)
    print(f"  {'#':<5} {'File':<{col}} {'Start (UTC)':<25} {'End (UTC)':<25} {'Dur':>10}  Method")
    print("  " + "-" * (col + 75))
    for i, c in enumerate(chunks, 1):
        print(
            f"  {i:<5} {c.path.name:<{col}} {fmt_ts(c.first_ts):<25} "
            f"{fmt_ts(c.last_ts):<25} {fmt_dur(c.duration_s):>10}  {c.ts_method}"
        )


def print_gap_report(gaps: list[Gap], threshold_s: float):
    print(f"\n{'=' * 70}")
    print(f"  GAP REPORT  (threshold: {threshold_s}s)")
    print(f"{'=' * 70}")

    if not gaps:
        print(f"\n  ✓  No gaps detected above {threshold_s}s threshold.\n")
        return

    print(f"\n  Found {len(gaps)} gap(s):\n")
    full_filled    = sum(1 for g in gaps if g.fillers and not g.partial)
    partial_filled = sum(1 for g in gaps if g.fillers and g.partial)
    unfilled       = sum(1 for g in gaps if not g.fillers)

    for i, g in enumerate(gaps, 1):
        if g.fillers and not g.partial:
            icon = "✓ FILLED"
        elif g.fillers and g.partial:
            last_filler  = max(g.fillers, key=lambda c: c.last_ts)
            first_filler = min(g.fillers, key=lambda c: c.first_ts)
            gap_before_s = (first_filler.first_ts - g.after.last_ts).total_seconds()
            gap_after_s  = (g.before.first_ts - last_filler.last_ts).total_seconds()
            parts = []
            if gap_before_s > 1:
                parts.append(f"{fmt_dur(gap_before_s)} before")
            if gap_after_s > 1:
                parts.append(f"{fmt_dur(gap_after_s)} after")
            icon = (f"~ PARTIAL ({g.coverage*100:.0f}% covered"
                    + (f"; still missing: {', '.join(parts)}" if parts else "")
                    + ")")
        else:
            icon = "⚠ UNFILLED"

        print(f"  [{i}] {icon}  — gap of {fmt_dur(g.gap_s)}")
        print(f"       After : {g.after.path.name}")
        print(f"               ends   at {fmt_ts(g.after.last_ts)}")
        print(f"       Before: {g.before.path.name}")
        print(f"               starts at {fmt_ts(g.before.first_ts)}")
        if g.fillers:
            print(f"       Fillers ({len(g.fillers)}):")
            for fc in sorted(g.fillers, key=lambda c: c.first_ts):
                print(f"         • {fc.path.name}  [{fmt_ts(fc.first_ts)} → {fmt_ts(fc.last_ts)}]")
        print()

    print(f"  Gaps fully filled    : {full_filled}")
    print(f"  Gaps partially filled: {partial_filled}")
    print(f"  Gaps unfilled        : {unfilled}")
    print(f"{'=' * 70}")


def print_summary(baseline: list[Chunk], fill: list[Chunk], gaps: list[Gap],
                  merged: list[Chunk], output_dir: Path, dry_run: bool,
                  min_coverage: float):
    full_filled    = [g for g in gaps if g.fillers and not g.partial]
    partial_filled = [g for g in gaps if g.fillers and g.partial]
    unfilled       = [g for g in gaps if not g.fillers]
    used_fill      = sum(len(g.fillers) for g in gaps)
    baseline_ts_ok = sum(1 for c in baseline if c.ts_ok)

    print(f"\n{'=' * 70}")
    print(f"  SUMMARY")
    print(f"{'=' * 70}")
    print(f"  Baseline chunks           : {len(baseline)}")
    print(f"    with readable timestamp : {baseline_ts_ok}")
    print(f"    unreadable timestamp    : {len(baseline) - baseline_ts_ok}  (still kept in output)")
    print(f"  Fill chunks available     : {len(fill)}")
    print(f"  Fill chunks used          : {used_fill}")
    print(f"  Gaps detected             : {len(gaps)}")
    print(f"  Gaps fully filled         : {len(full_filled)}")
    print(f"  Gaps partially filled     : {len(partial_filled)}  (>= {min_coverage*100:.0f}% coverage)")
    print(f"  Gaps still open           : {len(unfilled)}")
    print(f"  Total chunks in output    : {len(merged)}")
    print(f"  Output folder             : {output_dir}")
    print(f"  Mode                      : {'DRY RUN (no files written)' if dry_run else 'LIVE COPY'}")
    print(f"  Source files              : preserved (never deleted)")

    ts_ok = [c for c in merged if c.ts_ok]
    if len(ts_ok) >= 2:
        span_s = (ts_ok[-1].last_ts - ts_ok[0].first_ts).total_seconds()
        print(f"\n  Merged recording span     : {fmt_dur(span_s)}")
        print(f"    From : {fmt_ts(ts_ok[0].first_ts)}")
        print(f"    To   : {fmt_ts(ts_ok[-1].last_ts)}")

    if not unfilled and not partial_filled:
        status = "✓ COMPLETE (all gaps fully filled)"
    elif not unfilled and partial_filled:
        status = f"~ PARTIAL ({len(partial_filled)} gap(s) partially filled)"
    else:
        status = f"⚠ INCOMPLETE ({len(unfilled)} gap(s) remain empty)"

    print(f"\n  Status                    : {status}")
    print(f"{'=' * 70}\n")


# ============================================================
# EXCEL REPORT
# ============================================================

_HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_HEADER_FILL   = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
_TITLE_FONT    = Font(name="Arial", bold=True, size=14)
_LABEL_FONT    = Font(name="Arial", bold=True, size=10)
_BODY_FONT     = Font(name="Arial", size=10)
_STATUS_FILLS  = {
    "filled":   PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "partial":  PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "unfilled": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}
_STATUS_FONTS  = {
    "filled":   Font(name="Arial", color="006100", size=10, bold=True),
    "partial":  Font(name="Arial", color="9C6500", size=10, bold=True),
    "unfilled": Font(name="Arial", color="9C0006", size=10, bold=True),
}
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _style_header_row(ws, row: int, ncols: int):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER


def _autosize(ws, widths: dict[int, int]):
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_chunk_sheet(ws, chunks: list[Chunk], title: str):
    ws.title = title
    ws["A1"] = title
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:F1")

    headers = ["#", "File", "Start (UTC)", "End (UTC)", "Duration", "TS Method"]
    header_row = 3
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i, value=h)
    _style_header_row(ws, header_row, len(headers))

    r = header_row + 1
    for i, c in enumerate(chunks, start=1):
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=c.path.name)
        ws.cell(row=r, column=3, value=fmt_ts(c.first_ts))
        ws.cell(row=r, column=4, value=fmt_ts(c.last_ts))
        ws.cell(row=r, column=5, value=fmt_dur(c.duration_s))
        ws.cell(row=r, column=6, value=c.ts_method)
        for col in range(1, 7):
            cell = ws.cell(row=r, column=col)
            cell.font = _BODY_FONT
            cell.border = _BORDER
        r += 1

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    _autosize(ws, {1: 6, 2: 46, 3: 24, 4: 24, 5: 14, 6: 22})


def _write_gap_sheet(ws, gaps: list[Gap], threshold_s: float):
    ws.title = "Gap Report"
    ws["A1"] = f"Gap Report (threshold: {threshold_s}s)"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:H1")

    headers = ["#", "Status", "Gap Duration", "Coverage",
               "Chunk Before Gap", "Ends At",
               "Chunk After Gap", "Starts At", "Fillers Used"]
    header_row = 3
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i, value=h)
    _style_header_row(ws, header_row, len(headers))

    r = header_row + 1
    for i, g in enumerate(gaps, start=1):
        if g.fillers and not g.partial:
            status_key, status_label = "filled", "FILLED"
        elif g.fillers and g.partial:
            status_key, status_label = "partial", f"PARTIAL ({g.coverage*100:.0f}%)"
        else:
            status_key, status_label = "unfilled", "UNFILLED"

        filler_names = "; ".join(
            fc.path.name for fc in sorted(g.fillers, key=lambda c: c.first_ts)
        ) if g.fillers else "—"

        row_vals = [
            i, status_label, fmt_dur(g.gap_s),
            f"{g.coverage*100:.0f}%" if g.fillers else "0%",
            g.after.path.name, fmt_ts(g.after.last_ts),
            g.before.path.name, fmt_ts(g.before.first_ts),
            filler_names,
        ]
        for col, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = _BODY_FONT
            cell.border = _BORDER
            cell.alignment = Alignment(wrap_text=(col == 9), vertical="top")

        status_cell = ws.cell(row=r, column=2)
        status_cell.fill = _STATUS_FILLS[status_key]
        status_cell.font = _STATUS_FONTS[status_key]
        r += 1

    if not gaps:
        ws.cell(row=header_row + 1, column=1,
                 value=f"No gaps detected above {threshold_s}s threshold.").font = _BODY_FONT

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    _autosize(ws, {1: 5, 2: 18, 3: 14, 4: 10, 5: 40, 6: 22, 7: 40, 8: 22, 9: 60})


def _write_summary_sheet(ws, baseline: list[Chunk], fill: list[Chunk],
                         gaps: list[Gap], merged: list[Chunk],
                         output_dir: Path, dry_run: bool, min_coverage: float,
                         threshold_s: float, tolerance_s: float):
    ws.title = "Summary"
    ws["A1"] = "Gap Filler — Run Summary"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:B1")

    full_filled    = [g for g in gaps if g.fillers and not g.partial]
    partial_filled = [g for g in gaps if g.fillers and g.partial]
    unfilled       = [g for g in gaps if not g.fillers]
    used_fill      = sum(len(g.fillers) for g in gaps)
    baseline_ts_ok = sum(1 for c in baseline if c.ts_ok)

    if not unfilled and not partial_filled:
        status = "COMPLETE — all gaps fully filled"
        status_key = "filled"
    elif not unfilled and partial_filled:
        status = f"PARTIAL — {len(partial_filled)} gap(s) partially filled"
        status_key = "partial"
    else:
        status = f"INCOMPLETE — {len(unfilled)} gap(s) remain empty"
        status_key = "unfilled"

    ts_ok = [c for c in merged if c.ts_ok]
    span = ""
    span_from = ""
    span_to = ""
    if len(ts_ok) >= 2:
        span_s = (ts_ok[-1].last_ts - ts_ok[0].first_ts).total_seconds()
        span = fmt_dur(span_s)
        span_from = fmt_ts(ts_ok[0].first_ts)
        span_to = fmt_ts(ts_ok[-1].last_ts)

    rows = [
        ("Run timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Baseline folder", str(BASELINE_DIR)),
        ("Fill folder", str(FILL_DIR)),
        ("Output folder", str(output_dir)),
        ("Gap threshold (s)", threshold_s),
        ("Match tolerance (s)", tolerance_s),
        ("Min partial coverage", f"{min_coverage*100:.0f}%"),
        ("Mode", "DRY RUN (no files written)" if dry_run else "LIVE COPY"),
        ("", ""),
        ("Baseline chunks", len(baseline)),
        ("Baseline with readable timestamp", baseline_ts_ok),
        ("Baseline with unreadable timestamp", len(baseline) - baseline_ts_ok),
        ("Fill chunks available", len(fill)),
        ("Fill chunks used", used_fill),
        ("Gaps detected", len(gaps)),
        ("Gaps fully filled", len(full_filled)),
        ("Gaps partially filled", len(partial_filled)),
        ("Gaps still open", len(unfilled)),
        ("Total chunks in output", len(merged)),
        ("", ""),
        ("Merged recording span", span),
        ("Span start (UTC)", span_from),
        ("Span end (UTC)", span_to),
        ("", ""),
        ("Overall status", status),
    ]

    r = 3
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = _LABEL_FONT
        vcell = ws.cell(row=r, column=2, value=value)
        vcell.font = _BODY_FONT
        if label == "Overall status":
            vcell.fill = _STATUS_FILLS[status_key]
            vcell.font = _STATUS_FONTS[status_key]
        r += 1

    _autosize(ws, {1: 30, 2: 60})


def export_to_excel(baseline: list[Chunk], fill: list[Chunk], gaps: list[Gap],
                    merged: list[Chunk], output_dir: Path, dry_run: bool,
                    min_coverage: float, threshold_s: float,
                    tolerance_s: float) -> Path:
    """Build a multi-sheet Excel workbook summarizing the gap-fill run."""
    wb = Workbook()

    summary_ws = wb.active
    _write_summary_sheet(summary_ws, baseline, fill, gaps, merged, output_dir,
                         dry_run, min_coverage, threshold_s, tolerance_s)

    gap_ws = wb.create_sheet("Gap Report")
    _write_gap_sheet(gap_ws, gaps, threshold_s)

    baseline_ws = wb.create_sheet("Baseline Chunks")
    _write_chunk_sheet(baseline_ws, baseline, "Baseline Chunks")

    fill_ws = wb.create_sheet("Fill Chunks")
    _write_chunk_sheet(fill_ws, fill, "Fill Chunks")

    merged_ws = wb.create_sheet("Merged Sequence")
    _write_chunk_sheet(merged_ws, merged, "Merged Sequence")

    output_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = output_dir / "gap_report.xlsx"
    wb.save(xlsx_path)
    return xlsx_path


# ENTRY POINT
def main():
    parser = argparse.ArgumentParser(
        description="Fill gaps in baseline .dat chunks using a fill folder."
    )
    parser.add_argument("--threshold",    type=float, default=GAP_THRESHOLD_SECONDS,
                        help=f"Gap threshold in seconds (default: {GAP_THRESHOLD_SECONDS})")
    parser.add_argument("--tolerance",    type=float, default=MATCH_TOLERANCE_SECONDS,
                        help=f"Timestamp match tolerance in seconds (default: {MATCH_TOLERANCE_SECONDS})")
    parser.add_argument("--min-coverage", type=float, default=PARTIAL_FILL_MIN_COVERAGE,
                        help=f"Min fraction of gap a partial fill must cover (default: {PARTIAL_FILL_MIN_COVERAGE})")
    parser.add_argument("--verbose",      action="store_true",
                        help="Print full timestamp tables for all chunks")
    parser.add_argument("--dry-run",      action="store_true",
                        help="Detect and match only — do not copy any files")
    parser.add_argument("--no-excel",     action="store_true",
                        help="Skip writing the gap_report.xlsx summary workbook")
    parser.add_argument("--excel-path",   type=str, default=None,
                        help="Custom path for the Excel report (default: <output_dir>/gap_report.xlsx)")
    args = parser.parse_args()

    threshold    = args.threshold
    tolerance    = args.tolerance
    min_coverage = args.min_coverage

    print(f"\n{'=' * 70}")
    print(f"  GAP FILLER — .dat Chunk Merger")
    print(f"{'=' * 70}")
    print(f"  Baseline     : {BASELINE_DIR}")
    print(f"  Fill dir     : {FILL_DIR}")
    print(f"  Output       : {OUTPUT_DIR}")
    print(f"  Threshold    : {threshold}s")
    print(f"  Tolerance    : {tolerance}s")
    print(f"  Min coverage : {min_coverage*100:.0f}% (partial fills below this are discarded)")
    print(f"  Dry run      : {args.dry_run}")
    print(f"{'=' * 70}\n")

    for label, folder in [("BASELINE", BASELINE_DIR), ("FILL", FILL_DIR)]:
        if not folder.exists() or not folder.is_dir():
            print(f"[ERROR] {label} folder not found: {folder}")
            sys.exit(1)

    print("[1/5] Scanning baseline folder...")
    baseline = scan_folder(BASELINE_DIR, "baseline")
    baseline_ts_ok = sum(1 for c in baseline if c.ts_ok)
    print(f"      {len(baseline)} baseline chunk(s) found "
          f"({baseline_ts_ok} with readable timestamps, "
          f"{len(baseline) - baseline_ts_ok} kept without one).\n")

    print("[2/5] Scanning fill folder...")
    fill = scan_folder(FILL_DIR, "fill")
    print(f"      {len(fill)} fill chunk(s) found.\n")

    if args.verbose:
        if baseline: print_chunk_table(baseline, "BASELINE CHUNKS")
        if fill:     print_chunk_table(fill,     "FILL CHUNKS")

    print("[3/5] Detecting gaps in baseline...")
    gaps = detect_gaps(baseline, threshold)
    print(f"      {len(gaps)} gap(s) detected above {threshold}s.\n")

    match_fillers(gaps, fill, tol_s=tolerance, min_coverage=min_coverage)

    print_gap_report(gaps, threshold)

    print("[4/5] Building merged sequence...")
    merged = build_merged_sequence(baseline, gaps)

    copy_merged(merged, OUTPUT_DIR, dry_run=args.dry_run)

    print_summary(baseline, fill, gaps, merged, OUTPUT_DIR,
                  dry_run=args.dry_run, min_coverage=min_coverage)

    if not args.no_excel:
        print("[5/5] Writing Excel report...")
        excel_dir = Path(args.excel_path).parent if args.excel_path else OUTPUT_DIR
        xlsx_path = export_to_excel(
            baseline, fill, gaps, merged, OUTPUT_DIR,
            dry_run=args.dry_run, min_coverage=min_coverage,
            threshold_s=threshold, tolerance_s=tolerance,
        )
        if args.excel_path:
            custom_path = Path(args.excel_path)
            custom_path.parent.mkdir(parents=True, exist_ok=True)
            if custom_path != xlsx_path:
                shutil.move(xlsx_path, custom_path)
                xlsx_path = custom_path
        print(f"      Excel report written → {xlsx_path}\n")
    else:
        print("[5/5] Skipping Excel report (--no-excel).\n")


if __name__ == "__main__":
    main()