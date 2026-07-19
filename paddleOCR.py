from __future__ import annotations

import os
os.environ["PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK"] = "True"
os.environ["PADDLE_PDX_MODEL_SOURCE"] = "BOS"

import re
import struct
import shutil
import subprocess
import tempfile
import multiprocessing as mp
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter



INPUT_DIR   = Path(r"PATH_TO_INPUT_VIDEOS")  
OUTPUT_ROOT = Path(r"PATH_TO_OUTPUT_ROOT") 
IMAGES_ROOT = OUTPUT_ROOT / "images"

FFMPEG_PATH  = Path(r"PATH_TO_FFMPEG") 
FFPROBE_PATH = Path(r"PATH_TO_FFPROBE")  

MOVE_FILES = False

ROI_X = 0.00
ROI_Y = 0.00
ROI_W = 0.35
ROI_H = 0.18

VIDEO_EXTENSIONS = {".mp4", ".avi", ".mov", ".mkv", ".wmv", ".flv",
                    ".mpeg", ".mpg", ".dat"}

DISPLAY_TZ = timezone.utc

EXCEL_REPORT_PATH = OUTPUT_ROOT / "ocr_results.xlsx"

# GPUs to use — matches your nvidia-smi output
GPU_IDS = [0, 1]


# ----------------------------
# TIMESTAMP EXTRACTION
# ----------------------------

DAT_MARKERS = {
    b'\x80\x01\x00': "80 01 00 (H.264)",
    b'\x81\x01\x00': "81 01 00 (H.265)",
}


TS_MIN_US = 1_577_836_800 * 1_000_000
TS_MAX_US = 2_208_988_800 * 1_000_000


def _find_all_occurrences(data: bytes, pattern: bytes) -> list[int]:
    offsets, start = [], 0
    while True:
        pos = data.find(pattern, start)
        if pos == -1:
            break
        offsets.append(pos)
        start = pos + 1
    return offsets


def _read_ts_us(data: bytes, offset: int) -> int | None:
    if offset < 0 or offset + 8 > len(data):
        return None
    return struct.unpack_from('<Q', data, offset)[0]


def _is_valid_ts(ts_us: int) -> bool:
    return TS_MIN_US <= ts_us <= TS_MAX_US


def _extract_with_known_markers(data: bytes) -> list[dict]:
    results = []
    for marker_bytes, label in DAT_MARKERS.items():
        for pos in _find_all_occurrences(data, marker_bytes):
            ts_us = _read_ts_us(data, pos + 11)
            if ts_us and _is_valid_ts(ts_us):
                results.append({'offset': pos, 'ts_us': ts_us,
                                 'method': f"Known marker: {label}"})
    results.sort(key=lambda x: x['offset'])
    return results


def extract_timestamp_from_dat(dat_path: Path) -> tuple[datetime | None, str]:
    try:
        with open(dat_path, 'rb') as f:
            data = f.read()

        records = _extract_with_known_markers(data)
        method_used = "known markers"
        if not records:
            if records:
                by_method: dict[str, int] = {}
                for r in records:
                    by_method[r['method']] = by_method.get(r['method'], 0) + 1
                method_used = "; ".join(f"{m} ({c} hits)" for m, c in by_method.items())
            else:
                return None, "not found"

        middle = records[len(records) // 2]
        utc_dt = datetime.fromtimestamp(middle['ts_us'] / 1_000_000, tz=timezone.utc)
        return utc_dt, method_used

    except Exception as e:
        return None, f"error: {e}"

def safe_name(name: str) -> str:
    name = name.strip()
    name = re.sub(r'[<>:"/\\|?*]+', "_", name)
    name = re.sub(r"\s+", "_", name)
    return name or "unknown"


def ensure_unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem, suffix, parent = path.stem, path.suffix, path.parent
    counter = 2
    while True:
        candidate = parent / f"{stem}_{counter}{suffix}"
        if not candidate.exists():
            return candidate
        counter += 1


def get_video_codec(video_path: Path) -> str:
    """Returns 'hevc', 'h264', or 'unknown'."""
    result = subprocess.run(
        [str(FFPROBE_PATH), "-v", "error",
         "-select_streams", "v:0",
         "-show_entries", "stream=codec_name",
         "-of", "csv=p=0", str(video_path)],
        capture_output=True, text=True,
    )
    codec = result.stdout.strip().lower()
    if "hevc" in codec or "h265" in codec:
        return "hevc"
    if "h264" in codec or "avc" in codec:
        return "h264"
    return "unknown"

def get_middle_frame(video_path: Path, gpu_id: int = 0) -> np.ndarray:
    tmpdir = Path(tempfile.mkdtemp(prefix="ocr_frame_"))
    try:
        # Detect codec
        probe_codec = subprocess.run(
            [str(FFPROBE_PATH), "-v", "error",
             "-select_streams", "v:0",
             "-show_entries", "stream=codec_name",
             "-of", "csv=p=0", str(video_path)],
            capture_output=True, text=True,
        )
        codec = probe_codec.stdout.strip().lower()
        is_hevc = "hevc" in codec or "h265" in codec

        # Detect duration
        probe_dur = subprocess.run(
            [str(FFPROBE_PATH), "-v", "error",
             "-select_streams", "v:0",
             "-show_entries", "stream=duration",
             "-of", "csv=p=0", str(video_path)],
            capture_output=True, text=True,
        )
        duration = None
        try:
            val = float(probe_dur.stdout.strip().split()[0])
            if val > 0:
                duration = val
        except (ValueError, IndexError):
            pass

        print(f"  [FRAME] codec={codec} duration={f'{duration:.2f}s' if duration else 'unknown'}")

        def run_ffmpeg(post_seek: float | None) -> np.ndarray | None:
            out_path = tmpdir / "frame.jpg"
            if out_path.exists():
                out_path.unlink()

            seek = ["-ss", str(post_seek)] if post_seek is not None else []
            cmd = [
                str(FFMPEG_PATH), "-y",
                "-fflags", "+discardcorrupt+genpts",
                "-err_detect", "ignore_err",
                "-i", str(video_path),
            ] + seek + ["-vframes", "1", "-q:v", "2", str(out_path)]

            subprocess.run(cmd, capture_output=True)

            if out_path.exists() and out_path.stat().st_size > 10_000:  # real frame
                img = cv2.imread(str(out_path))
                if img is not None:
                    return img
            return None

        def roi_has_text(frame: np.ndarray) -> bool:
            """Quick check: does the ROI have enough contrast to contain text?"""
            h, w = frame.shape[:2]
            x1 = int(w * ROI_X);  y1 = int(h * ROI_Y)
            x2 = int(w * (ROI_X + ROI_W));  y2 = int(h * (ROI_Y + ROI_H))
            roi = frame[y1:y2, x1:x2]
            gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
            # Stddev: low = blank/solid, high = has content (text/overlay)
            _, stddev = cv2.meanStdDev(gray)
            return float(stddev[0][0]) > 20.0

        # Build seek positions to try
        if is_hevc:
            # Raw HEVC DAT: first frame always corrupt, try multiple offsets
            seek_positions = [2.0, 3.0, 5.0, 8.0, 10.0, 15.0]
            if duration:
                seek_positions = [duration * 0.5] + seek_positions
        else:
            # H.264: first frame usually fine
            seek_positions = [None, 1.0, 2.0, 3.0]
            if duration:
                seek_positions = [duration * 0.5, None, 1.0, 2.0]

        best_frame = None
        for seek in seek_positions:
            frame = run_ffmpeg(seek)
            if frame is None:
                continue
            if roi_has_text(frame):
                best_frame = frame
                print(f"  [FRAME] Good frame at seek={seek}s (stddev passed)")
                break
            elif best_frame is None:
                best_frame = frame  # keep as fallback even if no text detected

        if best_frame is None:
            raise RuntimeError(
                f"FFmpeg could not extract any frame from {video_path.name} "
                f"(codec={codec}, duration={duration})"
            )

        return best_frame
    finally:
        shutil.rmtree(str(tmpdir), ignore_errors=True)


# ----------------------------
# ROI + PREPROCESSING
# ----------------------------

def crop_top_left_roi(frame: np.ndarray) -> np.ndarray:
    h, w = frame.shape[:2]
    x1 = int(w * ROI_X);  y1 = int(h * ROI_Y)
    x2 = int(w * (ROI_X + ROI_W));  y2 = int(h * (ROI_Y + ROI_H))
    return frame[y1:y2, x1:x2].copy()


def preprocess_for_ocr(roi_bgr: np.ndarray) -> list[np.ndarray]:
    candidates: list[np.ndarray] = []

    # Existing full-ROI candidate
    raw_big = cv2.resize(
        roi_bgr,
        None,
        fx=2.5,
        fy=2.5,
        interpolation=cv2.INTER_CUBIC,
    )
    candidates.append(raw_big)

    gray = cv2.cvtColor(roi_bgr, cv2.COLOR_BGR2GRAY)
    gray = cv2.resize(
        gray,
        None,
        fx=2.5,
        fy=2.5,
        interpolation=cv2.INTER_CUBIC,
    )

    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    enhanced = clahe.apply(gray)
    candidates.append(cv2.cvtColor(enhanced, cv2.COLOR_GRAY2BGR))

    blur = cv2.GaussianBlur(enhanced, (3, 3), 0)
    thresholded = cv2.adaptiveThreshold(
        blur,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY,
        31,
        7,
    )
    candidates.append(cv2.cvtColor(thresholded, cv2.COLOR_GRAY2BGR))

    # Dedicated lower-left camera-label crop
    h, w = roi_bgr.shape[:2]

    channel_crop = roi_bgr[
        int(h * 0.40):int(h * 0.90),
        0:int(w * 0.38),
    ]

    # Padding prevents the text from being too close to the image boundary
    channel_crop = cv2.copyMakeBorder(
        channel_crop,
        25,
        25,
        25,
        25,
        cv2.BORDER_CONSTANT,
        value=(128, 128, 128),
    )

    channel_big = cv2.resize(
        channel_crop,
        None,
        fx=4.0,
        fy=4.0,
        interpolation=cv2.INTER_CUBIC,
    )
    candidates.append(channel_big)

    channel_gray = cv2.cvtColor(channel_big, cv2.COLOR_BGR2GRAY)
    channel_gray = clahe.apply(channel_gray)
    candidates.append(cv2.cvtColor(channel_gray, cv2.COLOR_GRAY2BGR))

    return candidates

def save_ocr_candidate_images(
    video_path: Path,
    images: list[np.ndarray],
) -> tuple[list[Path], list[Path]]:
    video_name = safe_name(video_path.stem)
    image_dir = IMAGES_ROOT / video_name
    image_dir.mkdir(parents=True, exist_ok=True)

    candidate_paths: list[Path] = []
    removable_paths: list[Path] = []

    for idx, img in enumerate(images, start=1):
        path = image_dir / f"{video_name}_roi_candidate_{idx}.png"

        success = cv2.imwrite(str(path), img)
        if not success:
            raise RuntimeError(f"Failed to save OCR candidate: {path}")

        candidate_paths.append(path)
        if idx > 3:
            removable_paths.append(path)

    return candidate_paths, removable_paths


# ----------------------------
# OCR HELPERS
# ----------------------------

_OCR_DT_PATTERNS = [
    r"\b(20\d{2})[/-](\d{1,2})[/-](\d{1,2})\s+(\d{1,2}):(\d{2}):(\d{2})\b",
]


def collect_useful_ocr_texts(data) -> list[str]:
    texts: list[str] = []
    if isinstance(data, dict):
        res = data.get("res", data)
        for block in res.get("parsing_res_list", []):
            if isinstance(block, dict):
                content = block.get("block_content")
                if isinstance(content, str) and content.strip():
                    texts.append(content.strip())
        for key in ("text", "rec_text", "content", "text_content", "markdown"):
            val = res.get(key)
            if isinstance(val, str) and val.strip():
                texts.append(val.strip())
    elif isinstance(data, list):
        for item in data:
            texts.extend(collect_useful_ocr_texts(item))
    elif isinstance(data, str):
        s = data.strip()
        if s:
            texts.append(s)

    clean: list[str] = []
    seen:  set[str]  = set()
    for t in texts:
        t = t.strip()
        if t and t not in seen:
            clean.append(t)
            seen.add(t)
    return clean


def normalize_ocr_text(text: str) -> str:
    text = text.upper().replace("|", "I")
    return re.sub(r"\s+", " ", text).strip()


def find_channel_label(texts: list[str]) -> str | None:
    joined = " ".join(normalize_ocr_text(t) for t in texts)
    for pat in [r"\bCAM([O0]\d{1,2})\b"]:
        m = re.search(pat, joined, flags=re.IGNORECASE)
        if m:
            digits = m.group(1).upper().replace("O", "0")  
            try:
                return f"CAM{int(digits):02d}"
            except ValueError:
                continue   
    return None


def find_ocr_datetime(texts: list[str]) -> datetime | None:
    joined = " ".join(t.strip() for t in texts if t.strip())
    for pat in _OCR_DT_PATTERNS:
        m = re.search(pat, joined)
        if not m:
            continue
        year, month, day, hour, minute, second = map(int, m.groups())
        try:
            return datetime(year, month, day, hour, minute, second,
                            tzinfo=DISPLAY_TZ)
        except ValueError:
            return None
    return None


def compare_dat_to_ocr(dat_timestamp: datetime | None,
                        ocr_timestamp: datetime | None) -> tuple[float | None, str]:
    if dat_timestamp is None and ocr_timestamp is None:
        return None, "no datetime found"
    if dat_timestamp is None:
        return None, "OCR only - no DAT reference"
    if ocr_timestamp is None:
        return None, "DAT only - OCR datetime missing"
    diff = abs((dat_timestamp - ocr_timestamp).total_seconds())
    status = ("PASS - DAT and OCR agree" if diff <= 10 else
              "CHECK - DAT and OCR close" if diff <= 60 else
              "REVIEW - DAT/OCR mismatch")
    return diff, status


def extract_overlay_from_image(ocr, image_path: Path,
                                debug: bool = False
                                ) -> tuple[str, datetime | None, list[str]]:
    results = ocr.predict(str(image_path))
    useful_texts: list[str] = []
    for idx, res in enumerate(results):
        data = res.json if hasattr(res, "json") else (
            res if isinstance(res, dict) else str(res))
        if debug:
            print(f"  [OCR DEBUG] Result #{idx}: {data}")
        useful_texts.extend(collect_useful_ocr_texts(data))

    clean_texts: list[str] = []
    seen: set[str] = set()
    for t in useful_texts:
        t = t.strip()
        if t and t not in seen:
            clean_texts.append(t)
            seen.add(t)

    channel_text = find_channel_label(clean_texts)
    channel      = safe_name(channel_text) if channel_text else "unknown_channel"
    ocr_dt       = find_ocr_datetime(clean_texts)
    return channel, ocr_dt, clean_texts


@dataclass
class VideoInfo:
    path:                   Path
    channel:                str
    ocr_timestamp:          datetime | None
    dat_timestamp:          datetime | None
    dat_ts_method:          str
    ocr_texts:              list[str]
    ocr_detected:           bool
    datetime_detected:      bool
    timestamp_diff_seconds: float | None
    timestamp_check:        str
    frame_ok:               bool
    error_msg:              str
    roi_path:               str
    seq:                    int = 0          
    output_path:            str = ""


# ----------------------------
# GPU WORKER
# ----------------------------

def gpu_worker(gpu_id: int,
               task_queue: mp.Queue,
               result_queue: mp.Queue) -> None:

    # Pin this process to one GPU
    os.environ["CUDA_VISIBLE_DEVICES"] = str(gpu_id)

    # Import here so each worker gets its own PaddlePaddle context
    from paddleocr import PaddleOCRVL

    print(f"[GPU {gpu_id}] Worker started (CUDA_VISIBLE_DEVICES={gpu_id})")

    ocr = PaddleOCRVL(
        use_layout_detection=True,
        use_doc_orientation_classify=False,
        use_doc_unwarping=False,
        device="gpu:0",
    )

    try:
        while True:
            item = task_queue.get()
            if item is None:       # sentinel — shut down
                break

            seq, video_path_str = item
            video_path = Path(video_path_str)
            print(f"[GPU {gpu_id}] [{seq}] Processing {video_path.name}")

            # ── Binary DAT timestamp ──────────────────────────────────────
            dat_timestamp, dat_ts_method = extract_timestamp_from_dat(video_path)

            # ── Frame + OCR ───────────────────────────────────────────────
            channel       = "unknown_channel"
            ocr_timestamp: datetime | None = None
            ocr_texts:     list[str]        = []
            ocr_detected       = False
            datetime_detected  = False
            frame_ok           = False
            error_msg          = ""
            roi_path_str       = ""

            try:
                frame = get_middle_frame(video_path, gpu_id=gpu_id)
                frame_ok = True
                roi      = crop_top_left_roi(frame)

                candidate_imgs = preprocess_for_ocr(roi)

                candidate_paths, removable_paths = save_ocr_candidate_images(
                    video_path,
                    candidate_imgs,
                )

                roi_path_str = " | ".join(str(p) for p in candidate_paths[:3])

                best_channel = "unknown_channel"
                best_ocr_dt: datetime | None = None
                best_texts: list[str] = []

                try:
                    for cp in candidate_paths:
                        ch_try, dt_try, texts_try = extract_overlay_from_image(
                            ocr,
                            cp,
                            debug=False,
                        )

                        for text in texts_try:
                            if text not in best_texts:
                                best_texts.append(text)

                        if best_channel == "unknown_channel" and ch_try != "unknown_channel":
                            best_channel = ch_try

                        if best_ocr_dt is None and dt_try is not None:
                            best_ocr_dt = dt_try

                        if best_channel != "unknown_channel" and best_ocr_dt is not None:
                            break

                finally:
                    for path in removable_paths:
                        try:
                            path.unlink(missing_ok=True)
                        except OSError as e:
                            print(f"  [WARNING] Could not delete temporary candidate {path}: {e}")

                channel           = best_channel
                ocr_timestamp     = best_ocr_dt
                ocr_texts         = best_texts
                ocr_detected      = channel != "unknown_channel"
                datetime_detected = ocr_timestamp is not None

                if not ocr_detected or not datetime_detected:
                    channel = "unknown_channel"

                print(f"[GPU {gpu_id}] [{seq}] "
                      f"dt={ocr_timestamp.strftime('%Y/%m/%d %H:%M:%S') if ocr_timestamp else 'N/A'} "
                      f"ch={channel}")

            except Exception as e:
                error_msg = str(e)
                print(f"[GPU {gpu_id}] [{seq}] ERROR: {e}")

            diff_seconds, timestamp_check = compare_dat_to_ocr(dat_timestamp,
                                                                 ocr_timestamp)

            info = VideoInfo(
                path                   = video_path,
                channel                = channel,
                ocr_timestamp          = ocr_timestamp,
                dat_timestamp          = dat_timestamp,
                dat_ts_method          = dat_ts_method,
                ocr_texts              = ocr_texts,
                ocr_detected           = ocr_detected,
                datetime_detected      = datetime_detected,
                timestamp_diff_seconds = diff_seconds,
                timestamp_check        = timestamp_check,
                frame_ok               = frame_ok,
                error_msg              = error_msg,
                roi_path               = roi_path_str,
                seq                    = seq,
            )
            result_queue.put(info)

    finally:
        ocr.close()
        print(f"[GPU {gpu_id}] Worker exiting")


# ----------------------------
# EXCEL REPORT
# ----------------------------

_COLUMNS = [
    ("No.",                    6),
    ("Filename",               40),
    ("Frame OK",               10),
    ("OCR Channel",            14),
    ("OCR Datetime",           22),
    ("DAT Reference UTC",      22),
    ("DAT vs OCR Diff Sec",    18),
    ("Timestamp Check",        28),
    ("Output File",            45),
    ("ROI Image Path",         45),
    ("Error",                  40),
]

_CLR_HEADER = "1F4E79"
_CLR_YES    = "E2EFDA"
_CLR_NO     = "FCE4D6"
_CLR_ERR    = "FFE0E0"
_CLR_ALT    = "F2F2F2"


def _make_border() -> Border:
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _hfont()   -> Font:      return Font(name="Arial", bold=True, color="FFFFFF", size=10)
def _bfont()   -> Font:      return Font(name="Arial", size=9)
def _center()  -> Alignment: return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _left()    -> Alignment: return Alignment(horizontal="left",   vertical="center", wrap_text=True)


def create_excel_report(output_path: Path):
    wb  = Workbook()
    ws  = wb.active
    ws.title = "OCR Results"
    hfill  = PatternFill("solid", fgColor=_CLR_HEADER)
    border = _make_border()
    for col_idx, (header, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _hfont(); cell.fill = hfill
        cell.alignment = _center(); cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws2 = wb.create_sheet("Summary")
    return wb, ws, ws2, 2


def append_row(ws, row_num: int, info: VideoInfo, seq: int):
    border = _make_border()
    ocr_ts_str = (info.ocr_timestamp.strftime("%Y-%m-%d %H:%M:%S UTC")
                  if info.ocr_timestamp else "")
    dat_ts_str = (info.dat_timestamp.strftime("%Y-%m-%d %H:%M:%S UTC")
                  if info.dat_timestamp else "")
    diff_str   = (round(info.timestamp_diff_seconds, 3)
                  if info.timestamp_diff_seconds is not None else "")

    if not info.frame_ok:
        bg = _CLR_ERR
    elif info.ocr_detected and info.datetime_detected:
        bg = _CLR_YES
    elif info.ocr_detected or info.datetime_detected:
        bg = _CLR_NO
    else:
        bg = _CLR_ALT if row_num % 2 == 0 else "FFFFFF"

    row_fill = PatternFill("solid", fgColor=bg)
    values = [seq, info.path.name,
              "Yes" if info.frame_ok else "No",
              info.channel if info.ocr_detected else "",
              ocr_ts_str, dat_ts_str, diff_str,
              info.timestamp_check, info.output_path,
              info.roi_path, info.error_msg]

    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = _bfont(); cell.fill = row_fill
        cell.border = border
        cell.alignment = (_center() if col_idx in (1, 3, 4, 7) else _left())
    ws.row_dimensions[row_num].height = 16


def write_summary_sheet(ws2, infos: list[VideoInfo]):
    hfill  = PatternFill("solid", fgColor=_CLR_HEADER)
    border = _make_border()
    total        = len(infos)
    frame_ok     = sum(1 for i in infos if i.frame_ok)
    ocr_detected = sum(1 for i in infos if i.ocr_detected)
    ocr_ts_found = sum(1 for i in infos if i.ocr_timestamp is not None)
    channels: dict[str, int] = {}
    for i in infos:
        channels.setdefault(i.channel, 0)
        channels[i.channel] += 1

    ws2.merge_cells("A1:C1")
    t = ws2["A1"]
    t.value = "OCR Pipeline — Summary"
    t.font  = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    t.fill  = hfill; t.alignment = _center()
    ws2.row_dimensions[1].height = 24

    stats = [
        ("Total files processed",    total),
        ("Frames extracted OK",      frame_ok),
        ("Frame extraction failed",  total - frame_ok),
        ("OCR label detected",       ocr_detected),
        ("OCR label NOT detected",   total - ocr_detected),
        ("Detection rate",           "=IFERROR(B5/B2,0)"),
        ("OCR datetimes found",      ocr_ts_found),
        ("OCR datetime missing",     total - ocr_ts_found),
    ]

    for cell_ref in ("A3", "B3"):
        c = ws2[cell_ref]
        c.value     = "Metric" if cell_ref == "A3" else "Value"
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = hfill; c.alignment = _center(); c.border = border

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 30

    for r_off, (metric, value) in enumerate(stats, start=4):
        ca = ws2.cell(row=r_off, column=1, value=metric)
        cb = ws2.cell(row=r_off, column=2, value=value)
        for c in (ca, cb):
            c.font = _bfont(); c.alignment = _left(); c.border = border
            c.fill = PatternFill("solid", fgColor=(
                _CLR_ALT if r_off % 2 == 0 else "FFFFFF"))
        if metric == "Detection rate":
            cb.number_format = "0.0%"

    start_row = len(stats) + 6
    for col, hdr in [(1, "Channel"), (2, "Count"), (3, "% of Total")]:
        c = ws2.cell(row=start_row, column=col, value=hdr)
        c.fill = hfill; c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.alignment = _center(); c.border = border

    for r_off, (ch, cnt) in enumerate(sorted(channels.items()), start=1):
        r    = start_row + r_off
        fill = PatternFill("solid", fgColor=(
            _CLR_ALT if r_off % 2 == 0 else "FFFFFF"))
        for col, val in [(1, ch), (2, cnt),
                         (3, f"=B{r}/B{start_row+len(channels)+1}" if total else 0)]:
            c = ws2.cell(row=r, column=col, value=val)
            c.font = _bfont(); c.alignment = _left()
            c.border = border; c.fill = fill
        ws2.cell(row=r, column=3).number_format = "0.0%"

    total_row = start_row + len(channels) + 1
    for col, val in [(1, "TOTAL"), (2, total), (3, 1.0)]:
        c = ws2.cell(row=total_row, column=col, value=val)
        c.font   = Font(name="Arial", bold=True, size=9)
        c.border = border
        c.fill   = PatternFill("solid", fgColor="D9D9D9")
    ws2.cell(row=total_row, column=3).number_format = "0.0%"



def copy_or_move_video(src: Path, dst: Path) -> Path:
    dst.parent.mkdir(parents=True, exist_ok=True)
    dst = ensure_unique_path(dst)
    if MOVE_FILES:
        shutil.move(str(src), str(dst))
    else:
        shutil.copy2(str(src), str(dst))
    return dst


def move_sorted_videos(infos: list[VideoInfo], ws):
    by_channel: dict[str, list[VideoInfo]] = {}
    for info in infos:
        by_channel.setdefault(info.channel, []).append(info)

    for channel, group in by_channel.items():
        channel_dir = OUTPUT_ROOT / channel
        channel_dir.mkdir(parents=True, exist_ok=True)

        def sort_key(x: VideoInfo):
            if channel == "unknown_channel":
                # Sort by the numeric chunk number just before the extension
                m = re.search(r'(\d+)$', x.path.stem)
                return int(m.group(1)) if m else float('inf')
            return (
                0 if x.ocr_timestamp is not None else 1,
                x.ocr_timestamp or datetime.max.replace(tzinfo=timezone.utc),
            )

        sorted_group = sorted(group, key=sort_key)

        print(f"\n[CHANNEL] {channel} — {len(sorted_group)} video(s)")
        for order, info in enumerate(sorted_group, start=1):
            destination = channel_dir / (f"{order:04d}_" + info.path.name)
            final_path  = copy_or_move_video(info.path, destination)
            info.output_path = str(final_path)
            print(f"  [{order:04d}] {info.path.name} -> {final_path.name}")

            # Update Output File cell in-place
            excel_row = info.seq + 1   # header = row 1
            ws.cell(row=excel_row, column=9, value=str(final_path))


def print_summary(infos: list[VideoInfo]):
    total        = len(infos)
    with_ts      = sum(1 for i in infos if i.ocr_timestamp is not None)
    frame_failed = sum(1 for i in infos if not i.frame_ok)
    detected     = sum(1 for i in infos if i.ocr_detected)
    channels     = sorted(set(i.channel for i in infos))

    print("\n" + "=" * 65)
    print("  SUMMARY")
    print("=" * 65)
    print(f"  Total files         : {total}")
    print(f"  Frame extracted OK  : {total - frame_failed}")
    print(f"  Frame extract failed: {frame_failed}")
    print(f"  OCR label detected  : {detected}  "
          f"({detected/total*100:.1f}%)" if total else "")
    print(f"  OCR datetime found  : {with_ts}")
    print(f"  OCR datetime missing: {total - with_ts}")
    print(f"\n  {'Channel':<25} {'Files':>6}")
    print("  " + "-" * 35)
    for ch in channels:
        count = sum(1 for i in infos if i.channel == ch)
        flag  = "  ✓" if ch != "unknown_channel" else ""
        print(f"  {ch:<25} {count:>6}{flag}")
    print("=" * 65)


def main():
    for binary in (FFMPEG_PATH, FFPROBE_PATH):
        if not binary.exists():
            print(f"[ERROR] FFmpeg binary not found: {binary}")
            return

    if not INPUT_DIR.exists():
        print(f"[ERROR] Input directory does not exist: {INPUT_DIR}")
        return

    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    IMAGES_ROOT.mkdir(parents=True, exist_ok=True)

    wb, ws, ws2, next_row = create_excel_report(EXCEL_REPORT_PATH)
    wb.save(str(EXCEL_REPORT_PATH))
    print(f"[INFO] Excel report : {EXCEL_REPORT_PATH}")
    print(f"[INFO] GPUs in use  : {GPU_IDS}")

    video_files = sorted(
        p for p in INPUT_DIR.iterdir()
        if p.is_file() and p.suffix.lower() in VIDEO_EXTENSIONS
    )
    if not video_files:
        print(f"[INFO] No video files found in: {INPUT_DIR}")
        return

    total = len(video_files)
    print(f"[INFO] Found {total} video(s) — distributing across {len(GPU_IDS)} GPUs")

    ctx         = mp.get_context("spawn")  
    task_queue  = ctx.Queue(maxsize=len(GPU_IDS) * 4)
    result_queue = ctx.Queue()

    workers = []
    for gpu_id in GPU_IDS:
        p = ctx.Process(
            target=gpu_worker,
            args=(gpu_id, task_queue, result_queue),
            daemon=True,
        )
        p.start()
        workers.append(p)

    def feed_tasks():
        for seq, vp in enumerate(video_files, start=1):
            task_queue.put((seq, str(vp)))
        # One sentinel per worker
        for _ in GPU_IDS:
            task_queue.put(None)

    import threading
    feeder = threading.Thread(target=feed_tasks, daemon=True)
    feeder.start()


    buffer: dict[int, VideoInfo] = {}
    infos:  list[VideoInfo]      = []
    next_expected = 1
    received      = 0

    while received < total:
        info = result_queue.get()
        received += 1
        buffer[info.seq] = info


        while next_expected in buffer:
            ready = buffer.pop(next_expected)
            infos.append(ready)
            append_row(ws, next_row, ready, seq=ready.seq)
            next_row += 1

            if next_expected % 10 == 0:
                wb.save(str(EXCEL_REPORT_PATH))
                print(f"[XLS] Progress saved ({next_expected}/{total})")

            next_expected += 1

        print(f"[PROGRESS] {received}/{total} completed")

    feeder.join()
    for p in workers:
        p.join()

    print(f"\n[INFO] Sorting and {'moving' if MOVE_FILES else 'copying'} videos...")
    move_sorted_videos(infos, ws)

    write_summary_sheet(ws2, infos)
    ws.auto_filter.ref = ws.dimensions
    wb.save(str(EXCEL_REPORT_PATH))
    print(f"\n[INFO] Excel report saved: {EXCEL_REPORT_PATH}")

    print_summary(infos)
    print("\n[DONE]")


if __name__ == "__main__":
    main()